import streamlit as st
import pandas as pd
import openpyxl
import os
import re
import io
import zipfile

def parse_variant(name):
    if ' - ' in name:
        parts = name.split(' - ')
        base_name = parts[0].strip()
        variant = parts[1]
        if ' / ' in variant:
            v_parts = variant.split(' / ')
            return base_name, v_parts[0].strip(), v_parts[1].strip()
    return name, None, None

def get_base_sku(sku):
    if pd.isna(sku):
        return None
    parts = str(sku).split('-')
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return str(sku)

def process_excel(csv_file, template_file, csv_filename):
    df = pd.read_excel(csv_file)

    # クライアント名はCSVファイル名のアンダースコア前の部分から取得
    # 例: "ぱるぷーら_2026-04.xlsx" → "ぱるぷーら"
    stem = os.path.splitext(csv_filename)[0]
    client_name = stem.split('_')[0]

    df['Base SKU'] = df['Lineitem sku'].apply(get_base_sku)
    df[['Product Base Name', 'Color', 'Size']] = df['Lineitem name'].apply(lambda x: pd.Series(parse_variant(x)))
    df = df.dropna(subset=['Color', 'Size'])

    products = df['Base SKU'].unique()
    output_files = []

    for sku in products:
        if not sku: continue

        prod_df = df[df['Base SKU'] == sku].copy()
        prod_name = prod_df['Product Base Name'].iloc[0]

        size_order = {'S': 0, 'M': 1, 'L': 2, 'XL': 3, '2XL': 4, 'XXL': 4, '3XL': 5, 'XXXL': 5}
        prod_df['Size_Order'] = prod_df['Size'].map(size_order)
        prod_df = prod_df.dropna(subset=['Size_Order'])

        agg = prod_df.groupby(['Color', 'Size'])['Lineitem quantity'].sum().unstack(fill_value=0)

        # Load Template from BytesIO
        template_file.seek(0)
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active

        # SKU formatting（ファイル名用）
        sku_display = sku
        match = re.match(r'(G\d+)-(\d+)', str(sku))
        if match:
            part1 = match.group(1)
            part2 = int(match.group(2))
            sku_display = f"{part1}-{part2:02d}"

        # デザイン欄の画像をすべて削除
        ws._images.clear()

        # テンプレートの固定値をクリア（空欄出力）
        ws.cell(row=1, column=7).value = None   # 発注日
        ws.cell(row=4, column=1).value = None   # 納期
        ws.cell(row=7, column=1).value = None   # 担当
        ws.cell(row=7, column=5).value = None   # デザイン名
        ws.cell(row=7, column=7).value = None   # アイテム名
        ws.cell(row=10, column=1).value = None  # プリント方法
        ws.cell(row=10, column=3).value = None  # ボディ品番
        ws.cell(row=10, column=4).value = None  # 通常発注

        # クライアント名はCSVファイル名から設定
        ws.cell(row=7, column=3).value = client_name

        colors = list(agg.index)

        # テンプレートに事前入力されたボディカラー・数量をすべてクリア
        for i in range(8):
            col_num = 3 + i
            ws.cell(row=12, column=col_num).value = None
            for row_num in range(13, 19):
                ws.cell(row=row_num, column=col_num).value = None
            ws.cell(row=19, column=col_num).value = None

        for c_idx, color in enumerate(colors):
            col_num = 3 + c_idx
            ws.cell(row=12, column=col_num).value = color

            sizes_map = {'S': 13, 'M': 14, 'L': 15, 'XL': 16, '2XL': 17, 'XXL': 17, '3XL': 18, 'XXXL': 18}
            for size_name, qty in agg.loc[color].items():
                if size_name in sizes_map:
                    ws.cell(row=sizes_map[size_name], column=col_num).value = qty

            col_sum = agg.iloc[c_idx].sum()
            ws.cell(row=19, column=col_num).value = col_sum

        grand_total = agg.values.sum()
        ws.cell(row=20, column=3).value = grand_total
        
        # Save to memory
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        
        safe_prod_name = re.sub(r'[\\/*?:"<>|]', "", prod_name)
        file_name = f"{sku_display}　{safe_prod_name}.xlsx"
        output_files.append((file_name, out_buf.getvalue()))

    return output_files

st.title("発注書自動作成システム")
st.write("ShopifyのCSVデータ(xlsx)をアップロードするだけで、商品別の発注書を自動作成します。")

# Fixed template path
TEMPLATE_FILE_PATH = "G2828-02　きなこがでろーんTシャツ.xlsx"

csv_file = st.file_uploader("CSVデータ (xlsx) をアップロード", type=["xlsx"])

if not os.path.exists(TEMPLATE_FILE_PATH):
    st.error(f"テンプレートファイル '{TEMPLATE_FILE_PATH}' が見つかりません。プログラムと同じフォルダに配置してください。")

if csv_file and os.path.exists(TEMPLATE_FILE_PATH):
    if st.button("発注書を作成する"):
        with st.spinner("集計・作成中..."):
            try:
                # Open the local template file as a binary stream
                with open(TEMPLATE_FILE_PATH, "rb") as f:
                    template_content = io.BytesIO(f.read())
                
                results = process_excel(csv_file, template_content, csv_file.name)
                
                # Create ZIP
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w") as zf:
                    for filename, content in results:
                        zf.writestr(filename, content)
                
                st.success(f"集計完了！ {len(results)} 商品分のファイルを生成しました。")
                st.download_button(
                    label="作成した発注書をダウンロード (ZIP形式)",
                    data=zip_buf.getvalue(),
                    file_name="generated_orders.zip",
                    mime="application/zip"
                )
            except Exception as e:
                st.error(f"処理中にエラーが発生しました: {e}")
