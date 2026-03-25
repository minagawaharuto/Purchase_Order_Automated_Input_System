import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re

def parse_variant(name):
    # Expected format: "Product Name - Color / Size"
    if ' - ' in name:
        parts = name.split(' - ')
        base_name = parts[0].strip()
        variant = parts[1]
        if ' / ' in variant:
            v_parts = variant.split(' / ')
            return base_name, v_parts[0].strip(), v_parts[1].strip()
    return name, None, None

def get_base_sku(sku):
    if pd.isna(sku):
        return None
    # Assuming SKU format G2828-002-09 -> G2828-002
    parts = str(sku).split('-')
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return str(sku)

def convert():
    input_file = 'CSVデータ.xlsx'
    template_file = 'G2828-02　きなこがでろーんTシャツ.xlsx'
    output_dir = 'output_orders'
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"Reading {input_file}...")
    df = pd.read_excel(input_file)
    
    # Pre-process
    df['Base SKU'] = df['Lineitem sku'].apply(get_base_sku)
    df[['Product Base Name', 'Color', 'Size']] = df['Lineitem name'].apply(lambda x: pd.Series(parse_variant(x)))
    
    # Filter out items without color/size (like simple products or misc items)
    df = df.dropna(subset=['Color', 'Size'])
    
    # Group by SKU
    products = df['Base SKU'].unique()
    
    for sku in products:
        if not sku: continue
        
        prod_df = df[df['Base SKU'] == sku]
        prod_name = prod_df['Product Base Name'].iloc[0]
        print(f"Processing SKU: {sku} ({prod_name})")
        
        # Aggregate
        # Size mapping for canonical order
        size_order = {'S': 0, 'M': 1, 'L': 2, 'XL': 3, '2XL': 4, 'XXL': 4, '3XL': 5, 'XXXL': 5}
        prod_df['Size_Order'] = prod_df['Size'].map(size_order)
        prod_df = prod_df.dropna(subset=['Size_Order'])
        
        agg = prod_df.groupby(['Color', 'Size'])['Lineitem quantity'].sum().unstack(fill_value=0)
        
        # Load Template
        wb = openpyxl.load_workbook(template_file)
        ws = wb['サンプル発注書']
        
        # Fill Header Info
        # Row 1, Col 3 (C1)
        # Convert G2828-002 to G2828-02 (removing one zero if it's 3-digit)
        sku_display = sku
        match = re.match(r'(G\d+)-(\d+)', str(sku))
        if match:
            part1 = match.group(1)
            part2 = int(match.group(2))
            sku_display = f"{part1}-{part2:02d}"
        
        ws.cell(row=1, column=3).value = sku_display
        
        # Row 7, Col 3 (C7) - Client Name
        ws.cell(row=7, column=3).value = 'ぱるぷーら' # Assuming fixed
        # Row 7, Col 5 (E7) - Design Name
        ws.cell(row=7, column=5).value = 'フロント'   # Assuming fixed
        # Row 7, Col 7 (G7) - Item Name
        ws.cell(row=7, column=7).value = 'Tシャツ'    # Assuming fixed
        
        # Fill Colors and Quantities
        colors = list(agg.index)
        # In template, colors are in C12, D12, E12... (Column 3, 4, 5...)
        for c_idx, color in enumerate(colors):
            col_num = 3 + c_idx
            ws.cell(row=12, column=col_num).value = color
            
            # Sizes are in Row 13 to 18 (S, M, L, XL, XXL, XXXL)
            sizes_map = {
                'S': 13,
                'M': 14,
                'L': 15,
                'XL': 16,
                '2XL': 17,
                'XXL': 17,
                '3XL': 18,
                'XXXL': 18
            }
            
            for size_name, qty in agg.loc[color].items():
                if size_name in sizes_map:
                    target_row = sizes_map[size_name]
                    ws.cell(row=target_row, column=col_num).value = qty

        # Formulas for totals (Row 19 and Col F/G/...) might already be in template, 
        # but let's make sure they are correct or recalculate if needed.
        # Template has totals in Row 19 (1-indexed).
        for c_idx in range(len(colors)):
            col_num = 3 + c_idx
            # Sum rows 13 to 18
            # ws.cell(row=19, column=col_num).value = f"=SUM({openpyxl.utils.get_column_letter(col_num)}13:{openpyxl.utils.get_column_letter(col_num)}18)"
            # Wait, let's just write the values to be safe
            col_sum = agg.iloc[c_idx].sum()
            ws.cell(row=19, column=col_num).value = col_sum
            
        # Grand Total in Row 20, Col 3 (C20 in template?)
        # Template Row 19 (Row 20 in 1-indexed) is "総枚数" (Grand Total)
        grand_total = agg.values.sum()
        ws.cell(row=20, column=3).value = grand_total
        
        # Save
        safe_prod_name = re.sub(r'[\\/*?:"<>|]', "", prod_name)
        output_name = f"{sku_display}　{safe_prod_name}.xlsx"
        wb.save(os.path.join(output_dir, output_name))
        print(f"Saved {output_name}")

if __name__ == "__main__":
    convert()
